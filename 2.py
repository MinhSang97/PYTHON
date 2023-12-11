import tkinter as tk
from tkinter import ttk
import pyodbc
import pandas as pd
from tkinter.filedialog import asksaveasfilename
from tkinter import messagebox
import configparser
from openpyxl import Workbook

# Tạo hoặc nạp cấu hình
config = configparser.ConfigParser()
config.read("config.ini")

# Kiểm tra nếu không có phần tử "Connection" trong cấu hình, tạo mới
if not config.has_section("Connection"):
    config.add_section("Connection")

# Lấy thông tin từ cấu hình (nếu có)
default_server = config.get("Connection", "Server", fallback="")
default_database = config.get("Connection", "Database", fallback="")
default_username = config.get("Connection", "Username", fallback="")
default_password = config.get("Connection", "Password", fallback="")


def test_connection():
    server_name = server_entry.get()
    database_name = db_entry.get()
    username = user_entry.get()
    password = password_entry.get()

    connection_string = f"DRIVER={{SQL Server}};SERVER={server_name};DATABASE={database_name};UID={username};PWD={password}"

    try:
        conn = pyodbc.connect(connection_string)
        messagebox.showinfo("Connection Test", "Connection successful!")
        conn.close()
    except pyodbc.Error as e:
        messagebox.showerror("Connection Test", f"Error: {str(e)}")


def save_configuration():
    server_name = server_entry.get()
    database_name = db_entry.get()
    username = user_entry.get()
    password = password_entry.get()

    # Ghi thông tin vào cấu hình
    config.set("Connection", "Server", server_name)
    config.set("Connection", "Database", database_name)
    config.set("Connection", "Username", username)
    config.set("Connection", "Password", password)
    with open("config.ini", "w") as configfile:
        config.write(configfile)

    # Hiển thị thông báo đã lưu thành công
    messagebox.showinfo("Save Configuration", "Configuration saved successfully!")


# Hàm để truy vấn
def execute_query():
    server_name = server_entry.get()
    database_name = db_entry.get()
    username = user_entry.get()
    password = password_entry.get()
    query = query_entry.get("1.0", tk.END).strip()

    connection_string = f"DRIVER={{SQL Server}};SERVER={server_name};DATABASE={database_name};UID={username};PWD={password}"

    try:
        conn = pyodbc.connect(connection_string)
        cursor = conn.cursor()

        cursor.execute(query)
        rows = cursor.fetchall()

        if len(rows) > 0:
            columns = [column[0] for column in cursor.description]
            columns = [col.strip() for col in columns]  # Xóa khoảng trắng trước và sau tên cột

            # Xóa dữ liệu trước đó trong Treeview
            result_tree.delete(*result_tree.get_children())

            # Hiển thị dữ liệu trên Treeview
            result_tree["columns"] = columns
            result_tree.column("#0", width=0, stretch=tk.NO)  # Ẩn cột đầu tiên
            for col in columns:
                result_tree.column(col, width=150, anchor=tk.W)
                result_tree.heading(col, text=col)
            for row in rows:
                result_tree.insert("", tk.END, values=row)
        else:
            # Xóa dữ liệu trước đó trong Treeview
            result_tree.delete(*result_tree.get_children())

            result_tree["columns"] = ("No results found",)
            result_tree.column("No results found", width=400)
            result_tree.heading("No results found", text="No results found.")

    except pyodbc.Error as e:
        # Xóa dữ liệu trước đó trong Treeview
        result_tree.delete(*result_tree.get_children())

        result_tree["columns"] = ("Error",)
        result_tree.column("Error", width=400)
        result_tree.heading("Error", text=f"Error: {str(e)}")



    # Hàm xuất kết quả thành file Excel
def export_to_excel():
    columns = result_tree["columns"]
    results = []
    for child in result_tree.get_children():
        values = result_tree.item(child)["values"]
        results.append(values)

    try:
        wb = Workbook()
        sheet = wb.active

        for col_num, column in enumerate(columns, 1):
            sheet.cell(row=1, column=col_num).value = column

        for row_num, row in enumerate(results, 2):
            for col_num, value in enumerate(row, 1):
                sheet.cell(row=row_num, column=col_num).value = value

        save_path = asksaveasfilename(defaultextension=".xlsx")
        if save_path:
            wb.save(save_path)
            messagebox.showinfo("Export Complete", "Data exported to Excel successfully!")
    except Exception as e:
        messagebox.showerror("Error", str(e))


    # Hàm thanh trượt ô kết quả truy vấn
def on_treeview_select(event):
    selected_item = result_tree.selection()[0]
    values = result_tree.item(selected_item, "values")
    result_text.delete("1.0", tk.END)
    result_text.insert(tk.END, ", ".join(values))
    result_text.see(tk.END)


#Tạo giao diện
root = tk.Tk()
root.title("SQL Query")
root.geometry("1920x1080")

#Tạo và đặt các label và entry fields cho thông tin kết nối
server_label = tk.Label(root, text="Server:")
server_label.pack()
server_entry = tk.Entry(root)
server_entry.pack()
server_entry.insert(0, default_server)

db_label = tk.Label(root, text="Database:")
db_label.pack()
db_entry = tk.Entry(root)
db_entry.pack()
db_entry.insert(0, default_database)

user_label = tk.Label(root, text="Username:")
user_label.pack()
user_entry = tk.Entry(root)
user_entry.pack()
user_entry.insert(0, default_username)

password_label = tk.Label(root, text="Password:")
password_label.pack()
password_entry = tk.Entry(root, show="*")
password_entry.pack()
password_entry.insert(0, default_password)

#Tạo nút kiểm tra kết nối
button_test_connection = tk.Button(root, text="Test Connection", command=test_connection)
button_test_connection.pack()

#Tạo nút lưu cấu hình
button_save_config = tk.Button(root, text="Save Configuration", command=save_configuration)
button_save_config.pack()

#Tạo label và text box cho truy vấn
query_label = tk.Label(root, text="Query:")
query_label.pack()
query_entry = tk.Text(root, height=10, width=200)
query_entry.pack()

#Tạo nút thực thi
execute_button = tk.Button(root, text="Execute", command=execute_query)
execute_button.pack()

# Tạo Treeview để hiển thị kết quả
result_label = tk.Label(root, text="Result:")
result_label.pack()


result_tree = ttk.Treeview(root)
result_tree.pack(fill=tk.BOTH, expand=True)

result_tree_scrollbar = ttk.Scrollbar(root, orient="vertical", command=result_tree.yview)
result_tree_scrollbar.pack(side=tk.RIGHT, fill=tk.Y)

result_tree.configure(yscrollcommand=result_tree_scrollbar.set)
result_tree.bind("<<TreeviewSelect>>", on_treeview_select)

# Thiết lập thanh trượt cho ô Treeview
result_tree.configure(yscrollcommand=result_tree_scrollbar.set)
result_tree_scrollbar.configure(command=result_tree.yview)

#Tạo nút xuất Excel
button_export = tk.Button(root, text="Export Excel", command=export_to_excel)
button_export.pack()

#Khởi chạy vòng lặp sự kiện Tkinter
root.mainloop()