import tkinter as tk
from tkinter import *
import cx_Oracle
from openpyxl import Workbook
from tkinter.filedialog import asksaveasfilename
from tkinter import messagebox

def test_connection():
    server_name = server_entry.get()
    port = port_entry.get()
    SID = service_entry.get()
    username = user_entry.get()
    password = password_entry.get()

    dsn = cx_Oracle.makedsn(server_name, port, service_name=SID)
    connection_string = f"{username}/{password}@{dsn}"

    try:
        conn = cx_Oracle.connect(connection_string)
        messagebox.showinfo("Connection Test", "Connection successful!")
        conn.close()
    except cx_Oracle.Error as e:
        messagebox.showerror("Connection Test", f"Error: {str(e)}")

# Hàm để truy vấn
def execute_query():
    server_name = server_entry.get()
    port = port_entry.get()
    SID = service_entry.get()
    username = user_entry.get()
    password = password_entry.get()
    query = query_entry.get("1.0", tk.END)

    dsn = cx_Oracle.makedsn(server_name, port, service_name=SID)
    connection_string = f"{username}/{password}@{dsn}"

    try:
        conn = cx_Oracle.connect(connection_string)
        cursor = conn.cursor()

        cursor.execute(query)
        rows = cursor.fetchall()

        result_text.delete("1.0", tk.END)  # Clear previous result

        if len(rows) > 0:
            columns = [column[0] for column in cursor.description]

            result_text.insert(tk.END, ", ".join(columns) + "\n")

            for row in rows:
                result_text.insert(tk.END, ", ".join(str(cell) for cell in row) + "\n")
        else:
            result_text.insert(tk.END, "No results found.")

        cursor.close()
        conn.close()
    except cx_Oracle.Error as e:
        result_text.delete("1.0", tk.END)  # Clear previous result
        result_text.insert(tk.END, f"Error: {str(e)}")

# Hàm xuất kết quả thành file Excel
def export_to_excel():
    columns = result_text.get("1.0", "1.end").strip().split(", ")
    results = [line.strip().split(", ") for line in result_text.get("2.0", tk.END).split("\n") if line.strip()]

    try:
        wb = Workbook()
        sheet = wb.active

        for col_num, column in enumerate(columns, 1):
            sheet.cell(row=1, column=col_num).value = column

        for row_num, row in enumerate(results, 2):
            for col_num, value in enumerate(row, 1):
                sheet.cell(row=row_num, column=col_num).value = value

        save_path = asksaveasfilename(defaultextension=".xlsx")
        if save_path:
            wb.save(save_path)
            messagebox.showinfo("Export Complete", "Data exported to Excel successfully!")
    except Exception as e:
        messagebox.showerror("Error", str(e))


# Create the main window
root = tk.Tk()
root.title("SQL Query")
root.geometry("1920x1080")

# Create labels and entry fields for connection information
server_label = tk.Label(root, text="Server:")
server_label.pack()
server_entry = tk.Entry(root)
server_entry.pack()

port_label = tk.Label(root, text="Port:")
port_label.pack()
port_entry = tk.Entry(root)
port_entry.pack()

service_label = tk.Label(root, text="SID:")
service_label.pack()
service_entry = tk.Entry(root)
service_entry.pack()

user_label = tk.Label(root, text="Username:")
user_label.pack()
user_entry = tk.Entry(root)
user_entry.pack()

password_label = tk.Label(root, text="Password:")
password_label.pack()
password_entry = tk.Entry(root, show="*")
password_entry.pack()

#Create test connection button
button_test_connection = tk.Button(root, text="Test Connection", command=test_connection)
button_test_connection.pack()

#Create query label and text box
query_label = tk.Label(root, text="Query:")
query_label.pack()
query_entry = tk.Text(root, height=20, width=200)
query_entry.pack()

#Create execute button
execute_button = tk.Button(root, text="Execute", command=execute_query)
execute_button.pack()

#Create result text box
result_label = tk.Label(root, text="Result:")
result_label.pack()
result_text = tk.Text(root, height=20, width=200)
result_text.pack()

#Create export button
button_export = tk.Button(root, text="Export Excel", command=export_to_excel)
button_export.pack()

#Start the Tkinter event loop
root.mainloop()


