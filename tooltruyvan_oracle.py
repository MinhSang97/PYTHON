
import tkinter as tk
from tkinter import ttk
import cx_Oracle
import pandas as pd
from tkinter.filedialog import asksaveasfilename
from tkinter import messagebox
import configparser
from openpyxl import Workbook
from PIL import Image, ImageTk
import ctypes
import sys
from configparser import ConfigParser

if getattr(sys, 'frozen', False):
    # Nếu đang chạy từ file .exe đã được build
    icon_path = 'D:/TEST/icon.ico'  # Đường dẫn tới file hình ảnh logo của bạn
    ctypes.windll.shell32.SetCurrentProcessExplicitAppUserModelID(icon_path)

# Tạo hoặc nạp cấu hình
config = configparser.ConfigParser()
config.read("config_oracle.ini")

# Kiểm tra nếu không có phần tử "Connection" trong cấu hình, tạo mới
if not config.has_section("Connection"):
    config.add_section("Connection")

# Lấy thông tin từ cấu hình (nếu có)
default_server = config.get("Connection", "Server", fallback="")
default_port = config.get("Connection", "Port", fallback="")
default_SID = config.get("Connection", "SID", fallback="")
default_username = config.get("Connection", "Username", fallback="")
default_password = config.get("Connection", "Password", fallback="")


def test_connection():
    server_name = server_entry.get()
    port = port_entry.get()
    SID = service_entry.get()
    username = user_entry.get()
    password = password_entry.get()

    dsn = cx_Oracle.makedsn(server_name, port, service_name=SID)
    connection_string = f"{username}/{password}@{dsn}"

    try:
        conn = cx_Oracle.connect(connection_string)
        messagebox.showinfo("Connection Test", "Connection successful!")
        conn.close()

        # Lưu thông tin vào file cấu hình

    except cx_Oracle.Error as e:
        messagebox.showerror("Connection Test", f"Error: {str(e)}")




def save_server_info(event=None):
    # Lấy thông tin từ các widget nhập liệu
    server_name = server_entry.get()
    port = port_entry.get()
    sid = service_entry.get()
    username = user_entry.get()
    password = password_entry.get()

    # Kiểm tra xem đã nhập đủ thông tin hay chưa
    if server_name and port and sid and username and password:
        # Lưu thông tin vào file cấu hình
        config = configparser.ConfigParser()
        config['SERVER'] = {
            'server_name': server_name,
            'port': port,
            'sid': sid,
            'username': username,
            'password': password
        }
        with open('config_oracle.ini', 'w') as configfile:
            config.write(configfile)
        # Hiển thị thông báo lưu thành công
        messagebox.showinfo("Success", "Server configuration saved.")
    else:
        # Hiển thị thông báo yêu cầu nhập đủ thông tin
        messagebox.showwarning("Incomplete Information", "Please enter all server information.")

# ...

def load_server_info():
    config = configparser.ConfigParser()
    config.read('config_oracle.ini')

    if 'SERVER' in config:
        server_name = config['SERVER'].get('server_name', '')
        port = config['SERVER'].get('port', '')
        sid = config['SERVER'].get('sid', '')
        username = config['SERVER'].get('username', '')
        password = config['SERVER'].get('password', '')

        server_entry.delete(0, 'end')
        server_entry.insert('end', server_name)
        port_entry.delete(0, 'end')
        port_entry.insert('end', port)
        service_entry.delete(0, 'end')
        service_entry.insert('end', sid)
        user_entry.delete(0, 'end')
        user_entry.insert('end', username)
        password_entry.delete(0, 'end')
        password_entry.insert('end', password)

        messagebox.showinfo("Success", "Server configuration loaded.")
    else:
        messagebox.showwarning("No Configuration", "No server configuration found.")

# Hàm để truy vấn
def execute_query():
    server_name = server_entry.get()
    port = port_entry.get()
    sid = service_entry.get()
    username = user_entry.get()
    password = password_entry.get()
    query = query_entry.get("1.0", tk.END)

    dsn = cx_Oracle.makedsn(server_name, port, service_name=sid)
    connection_string = f"{username}/{password}@{dsn}"

    try:
        conn = cx_Oracle.connect(connection_string)
        cursor = conn.cursor()

        cursor.execute(query)
        rows = cursor.fetchall()

        if len(rows) > 0:
            columns = [column[0] for column in cursor.description]
            columns = [col.strip() for col in columns]  # Xóa khoảng trắng trước và sau tên cột

            # Xóa dữ liệu trước đó trong Treeview
            result_tree.delete(*result_tree.get_children())

            # Hiển thị dữ liệu trên Treeview
            result_tree["columns"] = columns
            result_tree["displaycolumns"] = columns[0:]
            result_tree.heading("#0", text="", anchor=tk.W)
            result_tree.column("#0", width=0, stretch=False, anchor=tk.W)

            for col in columns:
                result_tree.column(col, anchor=tk.W, stretch=True)
                result_tree.heading(col, text=col)

            # Duyệt qua từng hàng và từng cột để hiển thị dữ liệu
            for row_num, row in enumerate(rows, 1):
                formatted_row = []
                for value in row:
                    # Chuyển giá trị thành chuỗi và kiểm tra nếu có dấu phẩy
                    if isinstance(value, float) or isinstance(value, int):
                        formatted_value = str(value)
                    else:
                        formatted_value = value
                    formatted_row.append(formatted_value)

                # Chèn hàng vào Treeview
                result_tree.insert("", tk.END, iid=row_num, values=formatted_row)

        else:
            # Xóa dữ liệu trước đó trong Treeview
            result_tree.delete(*result_tree.get_children())

            result_tree["columns"] = ("No results found",)
            result_tree.column("No results found", width=400)
            result_tree.heading("No results found", text="No results found.")

    except cx_Oracle.Error as e:
        # Xóa dữ liệu trước đó trong Treeview
        result_tree.delete(*result_tree.get_children())

        result_tree["columns"] = ("Error",)
        result_tree.column("Error", width=400)
        result_tree.heading("Error", text=f"Error: {str(e)}")


# Hàm xuất kết quả thành file Excel
def export_to_excel():
    columns = result_tree["columns"]
    results = []
    for child in result_tree.get_children():
        values = result_tree.item(child)["values"]
        results.append(values)

    try:
        wb = Workbook()
        sheet = wb.active

        for col_num, column in enumerate(columns, 1):
            sheet.cell(row=1, column=col_num).value = column

        for row_num, row in enumerate(results, 2):
            for col_num, value in enumerate(row, 1):
                sheet.cell(row=row_num, column=col_num).value = value

        save_path = asksaveasfilename(defaultextension=".xlsx")
        if save_path:
            wb.save(save_path)
            messagebox.showinfo("Export Complete", "Data exported to Excel successfully!")
    except Exception as e:
        messagebox.showerror("Error", str(e))

# Hàm thanh trượt ô kết quả truy vấn
def on_treeview_select(event):
    selected_item = result_tree.selection()[0]
    values = result_tree.item(selected_item, "values")
    result_text.delete("1.0", tk.END)
    result_text.insert(tk.END, ", ".join(values))
    result_text.see(tk.END)


def open_query_library_window():
    query_library_window = tk.Toplevel(root)
    query_library_window.title("Query Library")
    query_library_window.geometry("400x400")

    # Tạo label cho danh sách câu lệnh truy vấn mẫu
    query_library_label = tk.Label(query_library_window, text="Query Library:")
    query_library_label.pack()

    # Tạo frame để chứa nút cho từng câu lệnh truy vấn
    query_buttons_frame = tk.Frame(query_library_window)
    query_buttons_frame.pack(fill=tk.BOTH, expand=True)

    # Danh sách câu lệnh truy vấn mẫu
    query_library = [
        {
            "name": "Lệnh truy vấn dữ liệu đầu tháng 4",
            "query": "SELECT * FROM (Select * from A_Data_catalogue d left join BIZ_PUB_DATA_F_ENERGY_D c on d.Data_ID = c.Data_ID and cons_no like '%PA18%')  where RECEIVE_TIME like '%09-MAR-23%'"
        },
        {
            "name": "Lệnh truy vấn tất cả công tơ trên DCU",
            "query": "SELECT  Dcu.fcDcuAddr noDCU, Meter.fcMeterID noCôngTơ from Dcu INNER JOIN Meter ON  Dcu.fcDcuAddr=Meter.fcDcuID"
        },
        {
            "name": "Query 3",
            "query": "SELECT * FROM Table3"
        }
    ]

    # Tạo nút cho mỗi câu lệnh truy vấn
    for query_info in query_library:
        query_name = query_info["name"]
        query = query_info["query"]
        query_button = tk.Button(query_buttons_frame, text=query_name, command=lambda q=query: insert_query(q, query_entry))
        query_button.pack(fill=tk.X, padx=10, pady=5)


def insert_query(query, textbox):
    # Xóa nội dung hiện tại trong ô textbox
    textbox.delete(1.0, tk.END)

    # Chèn câu lệnh truy vấn vào ô textbox
    textbox.insert(tk.END, query)

# Tạo giao diện
root = tk.Tk()
root.title("HHM_ORACLE")

# Tạo một instance của Image từ file hình ảnh logo
logo_image = Image.open("icon.png")

# Tạo một instance của ImageTk từ Image
logo_image_tk = ImageTk.PhotoImage(logo_image)

# Thiết lập logo trong tiêu đề của cửa sổ
root.iconphoto(True, logo_image_tk)

#getting screen width and height of display
width= root.winfo_screenwidth()
height= root.winfo_screenheight()
root.geometry("%dx%d" % (width, height))
#root.geometry("1920x980")
root.attributes('-fullscreen',False)

# Tạo frame để chứa label và entry fields cho thông tin kết nối
connection_frame = tk.Frame(root)
connection_frame.pack(anchor=tk.W)

# Tạo và đặt label và entry fields cho thông tin kết nối
server_label = tk.Label(connection_frame, text="Server:")
server_label.grid(row=0, column=0)
server_entry = tk.Entry(connection_frame)
server_entry.grid(row=0, column=1)
server_entry.insert(0, default_server)

port_label = tk.Label(connection_frame, text="Port:")
port_label.grid(row=0, column=2)
port_entry = tk.Entry(connection_frame)
port_entry.grid(row=0, column=3)
port_entry.insert(0, default_port)

SID_label = tk.Label(connection_frame, text="SID:")
SID_label.grid(row=0, column=4)
service_entry = tk.Entry(connection_frame)
service_entry.grid(row=0, column=5)
service_entry.insert(0, default_username)

user_label = tk.Label(connection_frame, text="Username:")
user_label.grid(row=0, column=6)
user_entry = tk.Entry(connection_frame)
user_entry.grid(row=0, column=7)
user_entry.insert(0, default_username)




password_label = tk.Label(connection_frame, text="Password:")
password_label.grid(row=0, column=8)
password_entry = tk.Entry(connection_frame, show="*")
password_entry.grid(row=0, column=9)
password_entry.insert(0, default_password)



# Tạo nút kiểm tra kết nối, nút lưu cấu hình và nút mở cửa sổ chọn câu lệnh truy vấn mẫu
button_test_connection = tk.Button(connection_frame, text="Test Connection", command=test_connection)
button_test_connection.grid(row=0, column=10)

button_save_config = tk.Button(connection_frame, text="Save Configuration", command=save_server_info)
button_save_config.grid(row=0, column=11)

button_load_config = tk.Button(connection_frame, text="Load Configuration", command=load_server_info)
button_load_config.grid(row=0, column=12)

open_query_library_button = tk.Button(connection_frame, text="Open Query Library", command=open_query_library_window)
open_query_library_button.grid(row=0, column=13)

# Tạo label và text box cho truy vấn
query_label = tk.Label(root, text="Query:")
query_label.pack(anchor=tk.CENTER)
query_entry = tk.Text(root, height=10, width=250)
query_entry.pack(anchor=tk.W)

# Tạo nút thực thi
execute_button = tk.Button(root, text="Execute", command=execute_query)
execute_button.pack(anchor=tk.CENTER)

# Tạo Frame để chứa Treeview và thanh trượt
result_frame = tk.Frame(root)
result_frame.pack(fill=tk.BOTH, expand=True)

# Tạo thanh trượt dọc
result_tree_yscrollbar = ttk.Scrollbar(result_frame, orient="vertical")
result_tree_yscrollbar.pack(side=tk.RIGHT, fill=tk.Y)

# Tạo thanh trượt ngang
result_tree_xscrollbar = ttk.Scrollbar(result_frame, orient="horizontal")
result_tree_xscrollbar.pack(side=tk.BOTTOM, fill=tk.X)

# Tạo Treeview và kết nối với thanh trượt ngang và thanh trượt dọc
result_tree = ttk.Treeview(result_frame, yscrollcommand=result_tree_yscrollbar.set, xscrollcommand=result_tree_xscrollbar.set)
result_tree.pack(fill=tk.BOTH, expand=True)

# Thiết lập thanh trượt ngang và thanh trượt dọc
result_tree_yscrollbar.configure(command=result_tree.yview)
result_tree_xscrollbar.configure(command=result_tree.xview)

# Kết nối thanh trượt ngang và thanh trượt dọc với Treeview
result_tree.configure(yscrollcommand=result_tree_yscrollbar.set, xscrollcommand=result_tree_xscrollbar.set)


# Tạo nút xuất Excel
button_export = tk.Button(root, text="Export Excel", command=export_to_excel)
button_export.pack()

# Tạo Textbox để hiển thị giá trị khi chọn một dòng trong Treeview
result_text = tk.Text(root, height=1)
result_text.pack(fill=tk.X)

# Kết nối sự kiện chọn dòng trong Treeview với hàm on_treeview_select
result_tree.bind("<<TreeviewSelect>>", on_treeview_select)

# Khởi chạy vòng lặp sự kiện Tkinter
root.mainloop()
