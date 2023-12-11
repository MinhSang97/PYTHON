import tkinter as tk
from tkinter import ttk
import cx_Oracle
from configparser import ConfigParser
from openpyxl import Workbook
from tkinter.filedialog import asksaveasfilename
from tkinter import messagebox


def test_connection():
    server_name = server_entry.get()
    port = port_entry.get()
    SID = service_entry.get()
    username = user_entry.get()
    password = password_entry.get()

    dsn = cx_Oracle.makedsn(server_name, port, service_name=SID)
    connection_string = f"{username}/{password}@{dsn}"

    try:
        conn = cx_Oracle.connect(connection_string)
        messagebox.showinfo("Connection Test", "Connection successful!")
        conn.close()
        
        # Lưu thông tin vào file cấu hình
        save_server_info(server_name, port, SID, username, password)
    except cx_Oracle.Error as e:
        messagebox.showerror("Connection Test", f"Error: {str(e)}")


def save_server_info(server_name, port, SID, username, password):
    config = ConfigParser()
    config.read('config.ini')

    config['ServerInfo'] = {
        'ServerName': server_name,
        'Port': port,
        'SID': SID,
        'Username': username,
        'Password': password
    }

    with open('config.ini', 'w') as config_file:
        config.write(config_file)


def load_server_info():
    config = ConfigParser()
    config.read('config.ini')

    server_name = config.get('ServerInfo', 'ServerName', fallback='')
    port = config.get('ServerInfo', 'Port', fallback='')
    SID = config.get('ServerInfo', 'SID', fallback='')
    username = config.get('ServerInfo', 'Username', fallback='')
    password = config.get('ServerInfo', 'Password', fallback='')

    server_entry.delete(0, tk.END)
    server_entry.insert(0, server_name)
    port_entry.delete(0, tk.END)
    port_entry.insert(0, port)
    service_entry.delete(0, tk.END)
    service_entry.insert(0, SID)
    user_entry.delete(0, tk.END)
    user_entry.insert(0, username)
    password_entry.delete(0, tk.END)
    password_entry.insert(0, password)


# Hàm để truy vấn
def execute_query():
    server_name = server_entry.get()
    port = port_entry.get()
    SID = service_entry.get()
    username = user_entry.get()
    password = password_entry.get()
    query = query_entry.get("1.0", tk.END)

    dsn = cx_Oracle.makedsn(server_name, port, service_name=SID)
    connection_string = f"{username}/{password}@{dsn}"

    try:
        conn = cx_Oracle.connect(connection_string)
        cursor = conn.cursor()

        cursor.execute(query)
        rows = cursor.fetchall()

        result_tree.delete(*result_tree.get_children())  # Clear previous result

        if len(rows) > 0:
            columns = [column[0] for column in cursor.description]
            columns = [col.strip() for col in columns]  # Xóa khoảng trắng trước và sau tên cột

            # Hiển thị dữ liệu trên Treeview
            result_tree["columns"] = columns
            result_tree.column("#0", width=0, stretch=tk.NO)  # Ẩn cột đầu tiên
            for col in columns:
                result_tree.column(col, width=150, anchor=tk.W)
                result_tree.heading(col, text=col)
            for row in rows:
                result_tree.insert("", tk.END, values=row)
        else:
            result_tree["columns"] = ("No results found",)
            result_tree.column("No results found", width=400)
            result_tree.heading("No results found", text="No results found.")

    except cx_Oracle.Error as e:
        result_tree.delete(*result_tree.get_children())

        result_tree["columns"] = ("Error",)
        result_tree.column("Error", width=400)
        result_tree.heading("Error", text=f"Error: {str(e)}")


# Hàm xuất kết quả thành file Excel
def export_to_excel():
    columns = result_tree["columns"]
    results = []
    for child in result_tree.get_children():
        values = result_tree.item(child)["values"]
        results.append(values)

    try:
        wb = Workbook()
        sheet = wb.active

        for col_num, column in enumerate(columns, 1):
            sheet.cell(row=1, column=col_num).value = column

        for row_num, row in enumerate(results, 2):
            for col_num, value in enumerate(row, 1):
                sheet.cell(row=row_num, column=col_num).value = value

        save_path = asksaveasfilename(defaultextension=".xlsx")
        if save_path:
            wb.save(save_path)
            messagebox.showinfo("Export Complete", "Data exported to Excel successfully!")
    except Exception as e:
        messagebox.showerror("Error", str(e))


# Create the main window
root = tk.Tk()
root.title("SQL Query")
root.geometry("1920x1080")

# Create labels and entry fields for connection information
server_label = tk.Label(root, text="Server:")
server_label.pack()
server_entry = tk.Entry(root)
server_entry.pack()

port_label = tk.Label(root, text="Port:")
port_label.pack()
port_entry = tk.Entry(root)
port_entry.pack()

service_label = tk.Label(root, text="SID:")
service_label.pack()
service_entry = tk.Entry(root)
service_entry.pack()

user_label = tk.Label(root, text="Username:")
user_label.pack()
user_entry = tk.Entry(root)
user_entry.pack()

password_label = tk.Label(root, text="Password:")
password_label.pack()
password_entry = tk.Entry(root, show="*")
password_entry.pack()

# Load server info from config
load_server_info()

# Create test connection button
button_test_connection = tk.Button(root, text="Test Connection", command=test_connection)
button_test_connection.pack()

# Create save information button
button_save_info = tk.Button(root, text="Save Information", command=save_server_info)
button_save_info.pack()

# Create query label and text box
query_label = tk.Label(root, text="Query:")
query_label.pack()
query_entry = tk.Text(root, height=20, width=200)
query_entry.pack()

# Create execute button
execute_button = tk.Button(root, text="Execute", command=execute_query)
execute_button.pack()

# Tạo Treeview để hiển thị kết quả
result_label = tk.Label(root, text="Result:")
result_label.pack()
result_tree = ttk.Treeview(root)
result_tree.pack(fill=tk.BOTH, expand=True)

# Create export button
button_export = tk.Button(root, text="Export Excel", command=export_to_excel)
button_export.pack()

# Start the Tkinter event loop
root.mainloop()
