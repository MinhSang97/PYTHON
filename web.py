import tkinter as tk
from tkinter import filedialog
from openpyxl import load_workbook
import requests
import threading
import time
import datetime

# Hàm gửi yêu cầu lấy dữ liệu từ API
def get_data(cong_to, thoi_gian):
    # Gửi yêu cầu GET tới API với các tham số truyền vào
    url = f"http://192.168.30.252:5050/data?cong_to={cong_to}&thoi_gian={thoi_gian}"
    response = requests.get(url)

    # Kiểm tra mã phản hồi từ API
    if response.status_code == 200:
        data = response.json()
        # Trả về dữ liệu hoặc thực hiện các thao tác khác
        return data
    else:
        # Trả về None hoặc xử lý lỗi khác
        return None

# Hàm xử lý sự kiện khi nhấn nút "Chọn tệp"
def select_excel_file():
    file_path = filedialog.askopenfilename(filetypes=[("Excel Files", "*.xlsx")])
    if file_path:
        load_excel_file(file_path)

# Hàm xử lý tệp Excel được chọn
def load_excel_file(file_path):
    # Load workbook từ tệp Excel
    workbook = load_workbook(file_path)

    # Chọn sheet cần đọc dữ liệu
    sheet_name = 'Sheet1'
    sheet = workbook[sheet_name]

    # Đọc dữ liệu từ tệp Excel
    for row in sheet.iter_rows(min_row=2, values_only=True):
        cong_to = row[0]
        thoi_gian = row[1]

        # Định dạng lại giá trị thoi_gian
        if isinstance(thoi_gian, datetime.datetime):
            thoi_gian = thoi_gian.strftime("%d-%b-%y")
        elif isinstance(thoi_gian, str):
            try:
                thoi_gian = datetime.datetime.strptime(thoi_gian, "%Y-%m-%d %H:%M:%S").strftime("%d-%b-%y")
            except ValueError:
                print(f"Invalid datetime format: {thoi_gian}")
                continue
        else:
            print(f"Invalid datetime value: {thoi_gian}")
            continue

        # Gửi yêu cầu lấy dữ liệu từ API với các tham số từ tệp Excel
        threading.Thread(target=process_data, args=(cong_to, thoi_gian)).start()

    # Đóng workbook
    workbook.close()

# Tạo biến toàn cục để lưu trữ dữ liệu trả về từ API
data_buffer = []
data_lock = threading.Lock()  # Lock để đồng bộ hóa truy cập vào data_buffer

# Hàm cập nhật giao diện người dùng
def update_ui():
    while True:
        with data_lock:
            if data_buffer:
                data = data_buffer.pop(0)
                text_box.insert(tk.END, str(data) + "\n")
                text_box.see(tk.END)  # Cuộn xuống dòng cuối cùng

        # Nghỉ 0.1 giây trước khi kiểm tra lại dữ liệu trong buffer
        time.sleep(0.1)

# Hàm xử lý dữ liệu từ API và lưu vào data_buffer
def process_data(cong_to, thoi_gian):
    response_data = get_data(cong_to, thoi_gian)
    if response_data:
        # Cập nhật dữ liệu vào data_buffer
        with data_lock:
            data_buffer.append(response_data)
            # Gửi yêu cầu cập nhật giao diện người dùng
            root.event_generate("<<DataAvailable>>")

# Hàm xử lý sự kiện khi dữ liệu có sẵn để cập nhật giao diện người dùng
def handle_data_available(event):
    with data_lock:
        if data_buffer:
            data = data_buffer.pop(0)
            text_box.insert(tk.END, str(data) + "\n")
            text_box.see(tk.END)  # Cuộn xuống dòng cuối cùng

# Tạo giao diện đồ họa
root = tk.Tk()

# Tạo nút "Chọn tệp"
select_button = tk.Button(root, text="Chọn tệp", command=select_excel_file)
select_button.pack()



# Kết nối sự kiện "DataAvailable" với hàm xử lý tương ứng
root.bind("<<DataAvailable>>", handle_data_available)

# Bắt đầu luồng giao diện người dùng
threading.Thread(target=update_ui).start()

# Chạy ứng dụng
root.mainloop()


